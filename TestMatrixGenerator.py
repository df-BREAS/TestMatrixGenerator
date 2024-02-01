import os
import re
import sys
import csv
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def create_excel(filename, params):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Test Matrix"

    header = []
    for key in params.keys():
        header.append(key)
    
    ws.append(header)
    combos = get_combos(params)

    for row in combos:
        ws.append(row)

    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    wb.save(filename)
    wb.close()



def get_combos(params):
    result = []
    sublist_lens = []
    p = []
    for key in params:
        p.append(params[key])

    for sublist in p:
        sublist_lens.append(len(sublist))

    result2 = []
    recursive_helper(result2, p, sublist_lens, len(sublist_lens)*[0], len(sublist_lens), 0)
    return result2



def recursive_helper(result, p, sublist_lens, index_list, limit, ptr):
    if ptr is limit:
        result.append(return_combination_list(p, index_list))
    
    else:
        for i in range(sublist_lens[ptr]):
            index_list[ptr] = i
            recursive_helper(result, p, sublist_lens, index_list, limit, ptr + 1)



def return_combination_list(p, index_list):
    list = []
    i = 0
    for val in p:
        list.append(val[index_list[i]])
        i += 1
    
    return list

            

def get_params(filename):
    params = {}
    with open(filename) as f:
        for line in f:
            l = line.strip('\n').split(',')
            params[l[0]] = l[1:]
    
    return params



def main():
    if len(sys.argv) != 2:
        print("Usage: Parameters File Required")
        return

    parameters_file = sys.argv[1]
    params = get_params(parameters_file)
    filename = "Test Matrix.xlsx"
    create_excel(filename, params) 
    print("Successful creation of .xlsx file\n")
    


if __name__ == "__main__":
    main()